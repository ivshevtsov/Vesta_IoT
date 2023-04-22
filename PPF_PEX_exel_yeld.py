from openpyxl import load_workbook
import openpyxl
import numpy as np

#directory
HOME = 'DATA/PPF_PEX/csv'
File = 'gps_mism_pex'

wb = load_workbook(f'{HOME}/{File}.xlsx')
nsheet_table = 1
# get sheet names
sheetnames_table = wb.sheetnames
# get sheet data
sheet_table = wb[sheetnames_table[nsheet_table - 1]]

#print()

for col in range(13, 27+1):
    I=[]
    for rows in range(2,100):
        #print(rows)
        Value = float(sheet_table.cell(row=rows, column=col).value)
        I.append(Value)
    header = sheet_table.cell(row=1, column=col).value
    std = np.std(I)
    print(f'{header} - {std}')




