from openpyxl import load_workbook

HOME = '/Users/ivanshevtsov/PycharmProjects/Vesta_IoT/DATA/cadence_pads_exel'
FILE = 'PAD_NIGHTCALL.txt'
FILE_EXEL='PAD'
PITCH = 180
FILE_COMPARE = 'VESTA400_18_04_23'


file_error = open(f'{HOME}/Error.txt', "w")




wb = load_workbook(f'{HOME}/{FILE_EXEL}.xlsx')
wb_comp = load_workbook(f'{HOME}/{FILE_COMPARE}.xlsx')
# get sheet names
sheetnames = wb.sheetnames
sheetnames_comp = wb_comp.sheetnames
# get sheet data
nsheet=1
nsheet_comp=7
sheet = wb[sheetnames[nsheet - 1]]
sheet_comp = wb_comp[sheetnames_comp[nsheet_comp - 1]]



with open(f"{HOME}/{FILE}", "r") as file1:
    # итерация по строкам
    for line in file1:
        if line.strip().split()[1]=='Start':
            #print(line.strip().split())
            X_START = float(line.strip().split()[-2])
            Y_START = float(line.strip().split()[-1])
            #print(X_START)
            #print(Y_START)
#['Bump:', '"OUTP"', '4409.8', '5549.62']
with open(f"{HOME}/{FILE}", "r") as file1:
    # итерация по строкам
    for line in file1:
        if line.strip().split()[1]=='Bump:':
            #print(line.strip().split()[1:])
            X = float(line.strip().split()[-2])
            Y = float(line.strip().split()[-1])
            Name = str(line.strip().split()[2])[1:-1]
            X_C = round((X-X_START)/PITCH)
            Y_C = round((Y-Y_START)/PITCH)
            sheet.cell(row=62-Y_C, column=26-X_C).value = Name
            #print(f'{Name} X={X_C} Y={Y_C}')
            #print(f'X={X_C} Y={Y_C}')
            #print(Name)


wb.save(f'{HOME}/PAD.xlsx')


for row in range(2, 61+1):
    for column in range(2, 25+1):
        Chip_cadence = sheet.cell(row=row, column=column).value

        Chip_origin = sheet_comp.cell(row=row, column=column).value
        #print(f'{Chip_cadence}   {Chip_origin }')
        if (Chip_cadence!=Chip_origin) and Chip_cadence!='None' and Chip_origin!='no bump':
            print(f'{Chip_cadence}   {Chip_origin}')
            file_error.write(f'{Chip_cadence}   {Chip_origin}'+'\n')
file_error.close()