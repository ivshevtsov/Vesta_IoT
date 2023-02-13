from openpyxl import load_workbook
import openpyxl
from numpy import genfromtxt
import pandas as pd

HOME = 'DATA/Filters_corners'
File_main = 'corners_main_ppf_pex'
File_tune = 'corners_tune_ppf_pex'

#create corners
MODEL_FILE='Modelfile::/home/DATA/PDK/cadence_oa/t-n40-cm-sp-001-k3_2_0_2a_20140217_new/models/spectre/crn40lp_2d5_v2d0_2_shrink0d9_embedded_usage.scs'
TEST = 't Test::Vesta_Filters_40:Test_PPF_TOP_V1:2'



#PEX results
PEX_DATA = genfromtxt(f'{HOME}/{File_tune}.csv', delimiter=',', dtype='str')

wb = openpyxl.Workbook()
wb.save(f'{HOME}/{File_main}.xlsx')

nsheet=1
    # get sheet names
sheetnames = wb.sheetnames
    # get sheet data
sheet = wb[sheetnames[nsheet - 1]]


#create corners

sheet.cell(row=1, column=1).value = 'Corner'
sheet.cell(row=2, column=1).value = 'Enable'
sheet.cell(row=3, column=1).value = 'Temperature'
sheet.cell(row=4, column=1).value = 'VDD'

sheet.cell(row=5, column=1).value = 'B4'
sheet.cell(row=6, column=1).value = 'B3'
sheet.cell(row=7, column=1).value = 'B2'
sheet.cell(row=8, column=1).value = 'B1'
sheet.cell(row=9, column=1).value = 'B0'

sheet.cell(row=10, column=1).value = f'{MODEL_FILE}'
sheet.cell(row=11, column=1).value = f'{MODEL_FILE}'
sheet.cell(row=12, column=1).value = f'{MODEL_FILE}'

sheet.cell(row=13, column=1).value = f'{TEST}'

for i in range(len(PEX_DATA)-2):
    sheet.cell(row=1, column=i + 2).value = PEX_DATA[i + 2][0]
    sheet.cell(row=2, column=i + 2).value = 't'
    sheet.cell(row=3, column=i + 2).value = PEX_DATA[i + 2][2]
    sheet.cell(row=4, column=i + 2).value = PEX_DATA[i + 2][1]

    sheet.cell(row=5, column=i + 2).value = PEX_DATA[i + 2][9]
    sheet.cell(row=6, column=i + 2).value = PEX_DATA[i + 2][10]
    sheet.cell(row=7, column=i + 2).value = PEX_DATA[i + 2][11]
    sheet.cell(row=8, column=i + 2).value = PEX_DATA[i + 2][12]
    sheet.cell(row=9, column=i + 2).value = PEX_DATA[i + 2][13]

    sheet.cell(row=10, column=i + 2).value = 't '+PEX_DATA[i + 2][4]
    sheet.cell(row=11, column=i + 2).value ='t '+PEX_DATA[i + 2][5]
    sheet.cell(row=12, column=i + 2).value ='t '+PEX_DATA[i + 2][6]
    sheet.cell(row=13, column=i + 2).value = 't'


wb.save(f'{HOME}/{File_main}.xlsx')

data_xls = pd.read_excel(f'{HOME}/{File_main}.xlsx')
data_xls.to_csv(f'{HOME}/{File_main}.csv', encoding='utf-8', index=False)