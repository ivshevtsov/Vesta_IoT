from openpyxl import load_workbook

def get_bus_elements(str_pack, str_chip):
    #Dictionary
    DIC = {}

    #package string
    left_pack = str_pack.find('[')
    right_pack = str_pack.find(']')
    delta_pack = right_pack - left_pack

    l_num_pack=int(str_pack[right_pack - 1])
    h_num_pack = int(str_pack[left_pack+1:left_pack+1+delta_pack-3])

    #chip string
    left_chip = str_chip.find('[')
    right_chip = str_chip.find(']')
    delta_chip = right_chip - left_chip

    l_num_chip=int(str_chip[right_chip - 1])
    h_num_chip = int(str_chip[left_chip+1:left_chip+1+delta_chip-3])

    for i in range(l_num_pack, h_num_pack+1):
        str_0 = str_pack[:left_pack]+'['+f'{i}'+']' #package
        str_1 = str_chip[:left_chip] + f'{i+l_num_chip-l_num_pack}' #chip
        DIC[str_0]=str_1

    return DIC

#create dictionary from any input
def get_elements(str_pack, str_chip):

    DIC = {}
    if str_pack!=None:
        if str_pack[-1] == ']':
            DIC.update(get_bus_elements(str_pack, str_chip))
        else:
            DIC[str_pack] = str_chip
    return DIC


def rename_map_package(sheet, nrow, ncol, DIC):
    for row in range(2, nrow + 1):
        for column in range(2, ncol + 1):
            P_NAME = sheet.cell(row=row, column=column).value
            if P_NAME in DIC:
                sheet.cell(row=row, column=column).value = DIC[P_NAME]



HOME = 'DATA/Pinout_prove_change'
File = 'VestaTSTPinout_400(07.12.22)'
wb = load_workbook(f'{HOME}/{File}.xlsx')
# get sheet names
sheetnames = wb.sheetnames
# get sheet data

sheet_tb_gnss = wb[sheetnames[1 - 1]]
sheet_tb_test = wb[sheetnames[2 - 1]]
sheet_tb_modem = wb[sheetnames[3 - 1]]
sheet_tb_app_ui = wb[sheetnames[4 - 1]]
sheet_tb_app_ps = wb[sheetnames[5 - 1]]
sheet_map_pack = wb[sheetnames[6 - 1]]
sheet_map_chip = wb[sheetnames[7 - 1]]

DIC = {}

#generate set gnss
for i in range(3,40+1):
    str_pack =  sheet_tb_gnss.cell(row=i, column=1).value
    str_chip = sheet_tb_gnss.cell(row=i, column=8).value
    if str_chip!=None:
        DIC.update(get_elements(str_pack, str_chip))

#generate set tb_modem
for i in range(3,66+1):
    str_pack =  sheet_tb_modem.cell(row=i, column=1).value
    str_chip = sheet_tb_modem.cell(row=i, column=8).value
    if str_chip!=None:
        DIC.update(get_elements(str_pack, str_chip))

#generate set tb_app_ui
for i in range(3,67+1):
    str_pack =  sheet_tb_app_ui.cell(row=i, column=1).value
    str_chip = sheet_tb_app_ui.cell(row=i, column=5).value
    if str_chip!=None:
        DIC.update(get_elements(str_pack, str_chip))

#generate set tb_app_ps
for i in range(4,36+1):
    str_pack =  sheet_tb_app_ps.cell(row=i, column=1).value
    str_chip = sheet_tb_app_ps.cell(row=i, column=8).value
    if str_chip!=None:
        DIC.update(get_elements(str_pack, str_chip))


rename_map_package(sheet_map_pack, 21, 21, DIC)
wb.save(f'{HOME}/{File}_rename.xlsx')

'''


'''

print()





