from openpyxl import load_workbook

import re

def has_cyrillic(text):
    return bool(re.search('[а-яА-Я]', text))


def get_bus_elements(str_pack, str_chip):
    #Dictionary
    DIC = {}

    #package string
    left_pack = str_pack.find('[')
    right_pack = str_pack.find(']')
    delta_pack = right_pack - left_pack

    l_num_pack=int(str_pack[right_pack - 1])
    h_num_pack = int(str_pack[left_pack+1:left_pack+1+delta_pack-3])

    #chip string
    left_chip = str_chip.find('[')
    right_chip = str_chip.find(']')
    delta_chip = right_chip - left_chip

    l_num_chip=int(str_chip[right_chip - 1])
    h_num_chip = int(str_chip[left_chip+1:left_chip+1+delta_chip-3])

    for i in range(l_num_pack, h_num_pack+1):
        str_pack = str_pack[:left_pack]+f'{i}' #package
        str_chip = str_chip[:left_chip] + f'{i+l_num_chip-l_num_pack}' #chip
        DIC[str_chip]=str_pack

    return DIC

#create dictionary from any input
def get_elements(str_pack, str_chip):

    DIC = {}
    if (str_pack!=None) and (str_chip!=None):
        if str_pack[-1] == ']':
            DIC.update(get_bus_elements(str_pack, str_chip))
        else:
            LIST_CHIP = str_chip.replace(' ', '').split(',')
            if len(LIST_CHIP)>1:
                for i in LIST_CHIP:
                    DIC[i] = str_pack
            else:
                DIC[str_chip] = str_pack
    return DIC


def rename_map_chip(sheet, nrow, ncol, DIC):
    for row in range(2, nrow + 1):
        for column in range(2, ncol + 1):
            CHIP_NAME = sheet.cell(row=row, column=column).value
            if CHIP_NAME in DIC:
                sheet.cell(row=row, column=column).value = DIC[CHIP_NAME]



HOME = 'DATA/Pinout_prove_change'
File = 'VESTA400_09_02_23'
wb = load_workbook(f'{HOME}/{File}.xlsx')
# get sheet names
sheetnames = wb.sheetnames
# get sheet data

sheet_tb_gnss = wb[sheetnames[1 - 1]]
sheet_tb_test = wb[sheetnames[2 - 1]]
sheet_tb_modem = wb[sheetnames[3 - 1]]
sheet_tb_app_ui = wb[sheetnames[4 - 1]]
sheet_tb_app_ps = wb[sheetnames[5 - 1]]
sheet_map_pack = wb[sheetnames[6 - 1]]

sheet_map_chip = wb[sheetnames[7 - 1]]
#mirror
#sheet_map_chip = wb[sheetnames[8 - 1]]



DIC = {}

#generate DIC gnss
for i in range(3,44+1):
    str_pack =  sheet_tb_gnss.cell(row=i, column=1).value
    str_chip_d = sheet_tb_gnss.cell(row=i, column=8).value
    str_chip_a = sheet_tb_gnss.cell(row=i, column=9).value
    DIC.update(get_elements(str_pack, str_chip_a))
    DIC.update(get_elements(str_pack, str_chip_d))
    if (str_pack==None) and (str_chip_a!=None) and (str_chip_d!=None):
        DIC.update(get_elements(str_chip_d, str_chip_a))

for i in range(3,18+1):
    str_pack =  sheet_tb_test.cell(row=i, column=1).value
    str_chip_a = sheet_tb_test.cell(row=i, column=9).value
    DIC.update(get_elements(str_pack, str_chip_a))


#generate DIC modem
for i in range(3,68+1):
    str_pack =  sheet_tb_modem.cell(row=i, column=1).value
    str_chip_d = sheet_tb_modem.cell(row=i, column=8).value
    str_chip_a = sheet_tb_modem.cell(row=i, column=9).value
    DIC.update(get_elements(str_pack, str_chip_a))
    DIC.update(get_elements(str_pack, str_chip_d))


#generate set tb_app_ui
for i in range(3,70+1):
    str_pack =  sheet_tb_app_ui.cell(row=i, column=1).value
    str_chip = sheet_tb_app_ui.cell(row=i, column=5).value
    DIC.update(get_elements(str_pack, str_chip))

#generate set tb_app_ps
for i in range(4,39+1):
    str_pack =  sheet_tb_app_ps.cell(row=i, column=1).value
    str_chip_d = sheet_tb_app_ps.cell(row=i, column=8).value
    str_chip_a = sheet_tb_app_ps.cell(row=i, column=9).value
    DIC.update(get_elements(str_pack, str_chip_d))
    DIC.update(get_elements(str_pack, str_chip_a))

print(DIC)

rename_map_chip(sheet_map_chip, 62, 62, DIC)
wb.save(f'{HOME}/{File}_rename.xlsx')

#generate DIC test
