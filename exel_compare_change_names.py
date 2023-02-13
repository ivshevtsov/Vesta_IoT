from openpyxl import load_workbook

# get map set
def get_map_set(sheet, nrow, ncol):
    SET = set()
    for row in range(2, nrow + 1):
        for column in range(2, ncol + 1):
            V = sheet.cell(row=row, column=column).value
            V_str = str(V).strip()
            SET.add(V_str)
    return SET

#create dictionary from massive
def get_bus_elements(str_pack, str_chip):
    #Dictionary
    DIC = {}

    #package string
    left_pack = str_pack.find('[')
    right_pack = str_pack.find(']')
    delta_pack = right_pack - left_pack

    l_num_pack=int(str_pack[right_pack - 1])
    h_num_pack = int(str_pack[left_pack+1:left_pack+1+delta_pack-3])

    #chip string
    left_chip = str_chip.find('[')
    right_chip = str_chip.find(']')
    delta_chip = right_chip - left_chip

    l_num_chip=int(str_chip[right_chip - 1])
    h_num_chip = int(str_chip[left_chip+1:left_chip+1+delta_chip-3])

    for i in range(l_num_pack, h_num_pack+1):
        str_0 = str_pack[:left_pack]+f'{i}' #package
        str_1 = str_chip[:left_chip] + f'{i+l_num_chip-l_num_pack}' #chip
        DIC[str_0]=str_1

    return DIC

def get_one_elements_set(str):
    # Dictionary
    SET = set()

    # package string
    left_pack = str.find('[')
    right_pack = str.find(']')
    delta_pack = right_pack - left_pack
    if right_pack - left_pack > 2:
        l_num_pack = int(str[right_pack - 1])
        h_num_pack = int(str[left_pack + 1:left_pack + 1 + delta_pack - 3])
        for i in range(l_num_pack, h_num_pack + 1):
            if '_[' in str:
                str_0 = str[:left_pack] + f'{i}'  # element str
                SET.add(str_0)
            else:
                str_0 = str[:left_pack]  + f'{i}'  # element str
                SET.add(str_0)
    else:
        str_0=str[:left_pack] + f'{str[-2]}'
        SET.add(str_0)
    return SET

def get_one_elements_set_chip(str):
    # Dictionary
    SET = set()

    # package string
    left_pack = str.find('[')
    right_pack = str.find(']')
    delta_pack = right_pack - left_pack
    if right_pack - left_pack > 2:
        l_num_pack = int(str[right_pack - 1])
        h_num_pack = int(str[left_pack + 1:left_pack + 1 + delta_pack - 3])
        for i in range(l_num_pack, h_num_pack + 1):
            str_0 = str[:left_pack] + f'{i}'  # element str
            SET.add(str_0)
    else:
        str_0=str[:left_pack] + f'{str[-2]}'
        SET.add(str_0)
    return SET

#create dictionary from any input
def get_elements(str_pack, str_chip):

    DIC = {}

    if str_pack.find(']'):
        DIC.update(get_bus_elements(str_pack, str_chip))
    else:
        DIC[str_pack] = str_chip
    return DIC

def get_elements_set(string):
    SET = set()
    if string != None:
        if string[-1] ==']':
            SET.update(get_one_elements_set(string))
        else:

            #split and work with separator
            LIST = string.replace(" ", "").split(',')
            SET.update(LIST)
    return SET

def get_elements_set_chip(str):
    SET = set()
    if str != None:
        if str[-1] ==']':
            SET.update(get_one_elements_set_chip(str))
        else:
            #split and work with separator
            LIST = str.replace(' ','').split(',')
            SET.update(LIST)
    return SET


HOME = 'DATA/Pinout_prove_change'
File = 'VESTA400_09_02_23'
wb = load_workbook(f'{HOME}/{File}.xlsx')
# get sheet names
sheetnames = wb.sheetnames
# get sheet data
sheet_tb_gnss = wb[sheetnames[1 - 1]]
sheet_tb_test = wb[sheetnames[2 - 1]]
sheet_tb_modem = wb[sheetnames[3 - 1]]
sheet_tb_app_ui = wb[sheetnames[4 - 1]]
sheet_tb_app_ps = wb[sheetnames[5 - 1]]
sheet_map_pack = wb[sheetnames[6 - 1]]
sheet_map_chip = wb[sheetnames[7 - 1]]
#mirror
#sheet_map_chip = wb[sheetnames[8 - 1]]


SET_TABLE=set()
SET_MAP = set()

#generate set tb_gnss
for i in range(3,44+1):
    cell_left =  sheet_tb_gnss.cell(row=i, column=1).value
    cell_right = sheet_tb_gnss.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE.update(get_elements_set(cell_left))

#generate set tb_test
for i in range(3,18+1):
    cell_left =  sheet_tb_test.cell(row=i, column=1).value
    cell_right = sheet_tb_test.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE.update(get_elements_set(cell_left))

#generate set tb_modem
for i in range(3,68+1):
    cell_left =  sheet_tb_modem.cell(row=i, column=1).value
    cell_right = sheet_tb_modem.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE.update(get_elements_set(cell_left))


#generate set tb_app_ui
for i in range(3,70+1):
    cell_left =  sheet_tb_app_ui.cell(row=i, column=1).value
    cell_right = sheet_tb_app_ui.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE.update(get_elements_set(cell_left))

#generate set tb_app_ps
for i in range(4,39+1):
    cell_left =  sheet_tb_app_ps.cell(row=i, column=1).value
    cell_right = sheet_tb_app_ps.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE.update(get_elements_set(cell_left))


SET_MAP=get_map_set(sheet_map_pack, 21, 21)

print('Package errors')
print('Table - Map')
print(SET_TABLE-SET_MAP)
print('Map - Table')
print(SET_MAP-SET_TABLE)
print('#------------#')


##work with chip

SET_TABLE_CHIP=set()

#generate set tb_gnss
for i in range(3,44+1):
    cell_left_1 =  sheet_tb_gnss.cell(row=i, column=8).value
    cell_left_2 = sheet_tb_gnss.cell(row=i, column=9).value
    SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_1))
    SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_2))

#generate set tb_test
for i in range(3,18+1):
    cell_left =  sheet_tb_test.cell(row=i, column=9).value
    cell_right = sheet_tb_test.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left))

#generate set tb_modem
for i in range(3,68+1):
    cell_left_1 =  sheet_tb_modem.cell(row=i, column=9).value
    cell_left_2 = sheet_tb_modem.cell(row=i, column=8).value
    cell_right = sheet_tb_modem.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_1))
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_2))

#generate set tb_app_ui
for i in range(3,70+1):
    cell_left =  sheet_tb_app_ui.cell(row=i, column=5).value
    cell_right = sheet_tb_app_ui.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left))

#generate set tb_app_ps
for i in range(4,39+1):
    cell_left_1 =  sheet_tb_app_ps.cell(row=i, column=8).value
    cell_left_2 = sheet_tb_app_ps.cell(row=i, column=9).value
    cell_right = sheet_tb_app_ps.cell(row=i, column=2).value
    if cell_right!=None:
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_1))
        SET_TABLE_CHIP.update(get_elements_set_chip(cell_left_2))


SET_MAP_CHIP=get_map_set(sheet_map_chip, 62, 62)

print('Chip errors')
print('Table - Map')
print(SET_TABLE_CHIP-SET_MAP_CHIP)
print('Map - Table')
print(SET_MAP_CHIP-SET_TABLE_CHIP)
print('#------------#')


##---------------------------------------------------------
##---------------------------------------------------------
'''














print('RSV' in SET_TABLE_CHIP)
'''








