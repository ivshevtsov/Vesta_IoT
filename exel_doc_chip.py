from openpyxl import load_workbook
from docx import Document

def exel_dictionary_coord_chip(HOME,File, nsheet, nrow, ncol, pitch):
    #document
    doc = Document()
    table = doc.add_table(rows=1, cols=4)
    table.style = 'Table Grid'
    table.cell(0, 0).text = 'Номер вывода кристалла'
    table.cell(0, 1).text = 'Обозначение вывода кристалла'
    table.cell(0, 2).text = 'X\n мкм'
    table.cell(0, 3).text = 'Y\n мкм'

    # load document
    wb = load_workbook(f'{HOME}/{File}.xlsx')
    # get sheet names
    sheetnames = wb.sheetnames
    # get sheet data
    sheet = wb[sheetnames[nsheet - 1]]


    for row in range(2, nrow + 2):
        for column in range(2, ncol + 2):

            row_x = sheet.cell(row=row, column=1).value
            column_x = sheet.cell(row=1, column=column).value

            Value = sheet.cell(row=row, column=column).value

            Coordinate = f'{row_x}' + f'{column_x}'


            if (nrow-1)/2!=0:
                Coordinate_X = (-pitch * (ncol - 1) / 2) + (column - 2) * pitch
            else:
                Coordinate_X = (-pitch * (ncol - 0) / 2) + (column - 2) * pitch

            if (ncol - 1) / 2 != 0:
                Coordinate_Y = (pitch * (nrow - 1) / 2) - (row - 2) * pitch
            else:
                Coordinate_Y = (pitch * (nrow - 0) / 2) - (row - 2) * pitch

            if Value!= 'no bump':
                row_cells = table.add_row().cells
                row_cells[0].text = str(Coordinate)
                row_cells[1].text = str(Value)
                row_cells[2].text = str(Coordinate_X)
                row_cells[3].text = str(Coordinate_Y)

    doc.save(f'{HOME}/{File}.docx')

##------------------------------------------------##
##------------------------------------------------##

#directory
HOME = 'DATA/EXEL_convert'
File = 'CHIP_PADS'

#create doc file from exel (with coordinates from center)
#nrow/ncol - chip matrix size
#wating time 6-7 minutes

exel_dictionary_coord_chip(HOME, File=File, nsheet=1, nrow=61, ncol=61, pitch=162)
