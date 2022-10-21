from openpyxl import load_workbook
from docx import Document


def exel_dictionary(HOME,File, nsheet, nrow, ncol):
    # load document
    wb = load_workbook(f'{HOME}/{File}.xlsx')
    # get sheet names
    sheetnames = wb.sheetnames
    # get sheet data
    sheet = wb[sheetnames[nsheet - 1]]
    # list elements
    List_elements = []
    dic = {}
    with open(f'{HOME}/{File}.txt', "w") as file:
        for row in range(2, nrow + 2):
            for column in range(2, ncol + 2):
                row_x = sheet.cell(row=row, column=1).value
                column_x = sheet.cell(row=1, column=column).value
                Value = sheet.cell(row=row, column=column).value
                Coordinate = f'{row_x}' + f'{column_x}'
                String = Coordinate + ' ' +  f'{Value}'
                dic[Coordinate] = f'{Value}'
                # create list elements for compare
                List_elements.append(String)
                # print(String)
                file.write(f'{String} \n')
        file.close()
    return dic


HOME = 'DATA/EXEL_convert'
File = 'VestaPinout_400_1013'

#nrow/ncol - package matrix size
dic_package = exel_dictionary(HOME, File=File, nsheet=5, nrow=20, ncol=20)


#Create  common table
document = Document()
table = document.add_table(rows=1, cols=2)
table.style = 'Table Grid'
table.cell (0,0) .text = 'Номер вывода корпуса'
table.cell (0,1) .text = 'Обозначение вывода корпуса'


for key in dic_package:
    row_cells = table.add_row().cells
    row_cells[0].text = key
    row_cells[1].text = dic_package[key]

document.save(f'{HOME}/{File}.docx')