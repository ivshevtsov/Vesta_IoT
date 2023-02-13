from docx import Document
from PIL import Image, ImageDraw
import openpyxl
from openpyxl import load_workbook
import math

HOME = 'DATA/exel_generate_chip_pads'
File = 'TEST_BUMP_MAP_SCRIPTS'


Bumps = open(f'{HOME}/{File}.txt', 'r')
pitch=180

wb = openpyxl.Workbook()
wb.save(f'{HOME}/{File}.xlsx')
nsheet=1
    # get sheet names
sheetnames = wb.sheetnames
    # get sheet data
sheet = wb[sheetnames[nsheet - 1]]


for line in Bumps:
    line = line.strip()
    list_line = line.split(sep=' ')
    if list_line[0]=='Bump:':
        X = list_line[2]
        Y = list_line[3]

        Name = list_line[1]

        X_C = float(X)
        Y_C = float(Y)
        print(X + '' + Y)
        Xstart = -1643.78
        Ystart = -2130.55

        column =math.ceil(abs(X_C-Xstart)/pitch)+1
        row = 23-math.ceil(abs(Y_C-Ystart)/pitch)+1
        sheet.cell(row=row, column=column).value = Name



wb.save(f'{HOME}/{File}.xlsx')
