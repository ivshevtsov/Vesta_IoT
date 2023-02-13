from openpyxl import load_workbook


HOME = 'DATA/Pinout_prove_change'
File = 'VestaTSTPinout_400(15.12.22)'
wb = load_workbook(f'{HOME}/{File}.xlsx')
# get sheet names
sheetnames = wb.sheetnames
# get sheet data


sheet_map_chip = wb[sheetnames[7 - 1]]
sheet_map_chip_mirror = wb[sheetnames[8 - 1]]

nrow=61
ncol = 61


for row in range(2, nrow + 2):
    for column in range(2, ncol + 2):
        Value = sheet_map_chip.cell(row=row, column=column).value
        sheet_map_chip_mirror.cell(row=row, column=62-column+2).value = Value
wb.save(f'{HOME}/{File}.xlsx')