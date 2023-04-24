from openpyxl import load_workbook
from docx import Document
import openpyxl
import pandas as pd



def exel_dictionary_coord_chip(HOME,File, nsheet, nrow, ncol, pitch, name):
    #document

    #PinNumber
    #Padstack BPAD_500_400
    #XCoord
    #YCoord
    #Rotation
    #NetName
    wb_table = openpyxl.Workbook()
    wb_table.save(f'{HOME}/{name}_TABLE.xlsx')
    nsheet_table = 1
    # get sheet names
    sheetnames_table = wb_table.sheetnames
    # get sheet data
    sheet_table = wb_table[sheetnames_table[nsheet_table - 1]]
    sheet_table.cell(row=1, column=1).value = 'PinNumber'
    sheet_table.cell(row=1, column=2).value = 'Padstack'
    sheet_table.cell(row=1, column=3).value = 'XCoord'
    sheet_table.cell(row=1, column=4).value = 'YCoord'
    sheet_table.cell(row=1, column=5).value = 'NetName'


    # load document
    wb = load_workbook(f'{HOME}/{File}.xlsx')
    # get sheet names
    sheetnames = wb.sheetnames
    # get sheet data
    sheet = wb[sheetnames[nsheet - 1]]

    n_string = 1
    for row in range(2, nrow + 2):
        for column in range(2, ncol + 2):

            row_x = sheet.cell(row=row, column=1).value
            column_x = sheet.cell(row=1, column=column).value

            Value = sheet.cell(row=row, column=column).value

            Coordinate = f'{row_x}' + f'{column_x}'
            print(Coordinate)

            if (nrow-1)/2!=0:
                Coordinate_X = (-pitch * (ncol - 1) / 2) + (column - 2) * pitch
            else:
                Coordinate_X = (-pitch * (ncol - 0) / 2) + (column - 2) * pitch

            if (ncol - 1) / 2 != 0:
                Coordinate_Y = (pitch * (nrow - 1) / 2) - (row - 2) * pitch
            else:
                Coordinate_Y = (pitch * (nrow - 0) / 2) - (row - 2) * pitch

            if (Value!= 'no bump') and (Value!=None):
                n_string = n_string + 1
                sheet_table.cell(row=n_string, column=1).value = Coordinate
                sheet_table.cell(row=n_string, column=2).value = 'BPAD_500_400'
                sheet_table.cell(row=n_string, column=3).value = Coordinate_X
                sheet_table.cell(row=n_string, column=4).value = Coordinate_Y
                sheet_table.cell(row=n_string, column=5).value = Value.strip()

    wb_table.save(f'{HOME}/{name}_TABLE.xlsx')
    data_xls = pd.read_excel(f'{HOME}/{name}_TABLE.xlsx')
    data_xls.to_csv(f'{HOME}/{name}_TABLE.txt', encoding='utf-8', index=False, sep=' ')
##------------------------------------------------##
##------------------------------------------------##

#directory
HOME = 'DATA/Pinout_prove_change'
File_package = 'VESTA400_18_04_23'
File_chip =    'VESTA400_18_04_23_rename'

#exel_compare_change_names
#exel_chip_rename_lay
#exel_table_for_layouter

exel_dictionary_coord_chip(HOME, File=File_package, nsheet=6, nrow=20, ncol=20, pitch=800, name='package_18_04_23')
exel_dictionary_coord_chip(HOME, File=File_chip, nsheet=7, nrow=61, ncol=61, pitch=162, name='chip_18_04_23')


