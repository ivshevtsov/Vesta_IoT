from openpyxl import load_workbook
from docx import Document

##------------------------------------------------##
##------------------------------------------------##
def get_key(d, value):
    string=''
    for k, v in d.items():
        if v == value:
            string = string + ' ' +str(k)
    return string

def exel_dictionary(HOME,File, nsheet, nrow, ncol):
    sep = '	'
    # load document
    wb = load_workbook(f'{HOME}/{File}.xlsx')
    # get sheet names
    sheetnames = wb.sheetnames
    # get sheet data
    sheet = wb[sheetnames[nsheet - 1]]
    # list elements
    List_elements = []
    dic = {}
    with open(f'{HOME}/{File}.txt', "w") as file:
        for row in range(2, nrow + 1):
            for column in range(2, ncol + 1):
                row_x = sheet.cell(row=row, column=1).value
                column_x = sheet.cell(row=1, column=column).value
                Value = sheet.cell(row=row, column=column).value
                Coordinate = f'{row_x}' + f'{column_x}'
                String = Coordinate + sep +  f'{Value}'
                dic[Coordinate] = f'{Value}'
                # create list elements for compare
                List_elements.append(String)
                # print(String)
                file.write(f'{String} \n')
        file.close()
    return dic

##------------------------------------------------##
##------------------------------------------------##

#directory
HOME = 'DATA/EXEL_convert'
File = 'GNSS_CHIP'
#create txt files from exel
dic_package = exel_dictionary(HOME, File=File, nsheet=1, nrow=24, ncol=18)
