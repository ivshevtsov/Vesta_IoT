from openpyxl import load_workbook
from docx import Document

HOME = 'DATA/EXEL_convert'
File = 'Converter'

wb = load_workbook(f'{HOME}/{File}.xlsx')
nsheet=1
    # get sheet names
sheetnames = wb.sheetnames
    # get sheet data
sheet = wb[sheetnames[nsheet - 1]]





sheet.cell(row=2, column=2).value='Hello'

wb.save(f'{HOME}/{File}.xlsx')