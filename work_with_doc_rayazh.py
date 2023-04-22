from docx import Document
from PIL import Image, ImageDraw
from openpyxl import load_workbook
import math
import openpyxl


HOME = 'DATA/EXEL_convert'
File = '2023-02-06-nviot01_top-bump-locations-full-edited.txt'

HOME_EXEL = 'DATA/Pinout_prove_change'
File_exel = 'VESTA400_06_02_23_V1'



im = Image.new('RGB', (6000, 10000), 'white')
draw = ImageDraw.Draw(im)


Bumps = open(f'{HOME}/{File}', 'r')
D = 100
R = int(D/2)
pitch=180

#wb = openpyxl.Workbook()
#wb.save(f'{HOME}/CHIP_PADS.xlsx')

wb = load_workbook(f'{HOME_EXEL}/{File_exel}.xlsx')
nsheet=7
    # get sheet names
sheetnames = wb.sheetnames
    # get sheet data
sheet = wb[sheetnames[nsheet - 1]]


for line in Bumps:
    line = line.strip()
    list_line = line.split(sep=' ')
    if list_line[0]=='Bump:':
        X = list_line[3]
        Y = list_line[4]

        if len(list_line)>=6:
            Name = list_line[5]
        else:
            Name = 'NC'

        X_C = float(X)
        Y_C = float(Y)
        Xstart = 149.890
        Ystart = 149.620

        column =math.ceil(62-abs(X_C-Xstart)/pitch)
        row = math.ceil(62-abs(Y_C-Ystart)/pitch)
        sheet.cell(row=row, column=column).value = Name


        X_Ci = int(X_C)
        Y_Ci = int(Y_C)
        if Name != 'no_ball':
            draw.ellipse((X_Ci - R, Y_Ci - R, X_Ci + R, Y_Ci + R), fill=(1, 0, 0, 0))
            print(Name + ' ' + str(row) + ' ' + str(column))
        else:
            None

wb.save(f'{HOME_EXEL}/{File_exel}.xlsx')

im.show()
